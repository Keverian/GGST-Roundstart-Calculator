import os
from pathlib import Path
import openpyxl
from openpyxl import Workbook
import Character, os, Main, WebScrape
storePath = Path.cwd()/Path(r'characterFolder')
#Check if the character directory that store all sheets exist, if not make that folder
if not os.path.exists(storePath):
    storePath.mkdir()

nameList = WebScrape.getAllName()
#change CWD to character directory
os.chdir(str(storePath))

"""
#function that creates a workbook for a character
def createWorkbook(character: Character, CharacterList):
    if not os.path.exists(storePath):
        storePath.mkdir()
    wb = Workbook()
    ws = wb.active()
"""
def createWBTest(character: Character):
    if not os.path.exists(storePath):
        storePath.mkdir()
        os.chdir(str(storePath))

    wb = Workbook()
    ws = wb.active()
    ws.title = character.name








